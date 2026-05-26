from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from dateutil import parser as date_parser
from openpyxl import load_workbook

from .models import Property, Recipient


def _parse_bool(value: Any, default: bool = True) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y", "同意", "ok", "active"}


def _parse_datetime(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    if value is None or str(value).strip() == "":
        raise ValueError("updated_at is required")
    return date_parser.parse(str(value))


def _rows_from_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def _rows_from_json(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        data = data.get("properties") or data.get("recipients") or data.get("data") or []
    if not isinstance(data, list):
        raise ValueError(f"JSON root must be a list: {path}")
    return [dict(item) for item in data]


def _rows_from_xlsx(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    result: list[dict[str, Any]] = []
    for row in rows[1:]:
        item = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}
        if any(value not in (None, "") for value in item.values()):
            result.append(item)
    return result


def read_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _rows_from_csv(path)
    if suffix == ".json":
        return _rows_from_json(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _rows_from_xlsx(path)
    raise ValueError(f"Unsupported data file type: {path}")


def normalize_property(row: dict[str, Any]) -> Property:
    return Property(
        property_id=str(row.get("property_id") or row.get("id") or "").strip(),
        title=str(row.get("title") or row.get("name") or "").strip(),
        price=float(str(row.get("price") or 0).replace(",", "")),
        area=float(str(row.get("area") or row.get("sqm") or 0).replace(",", "")),
        address=str(row.get("address") or "").strip(),
        status=str(row.get("status") or "active").strip().lower(),
        updated_at=_parse_datetime(row.get("updated_at") or row.get("updated") or row.get("modified_at")),
        detail_url=str(row.get("detail_url") or row.get("url") or "").strip(),
        brochure_url=str(row.get("brochure_url") or row.get("document_url") or "").strip() or None,
        station=str(row.get("station") or "").strip() or None,
        layout=str(row.get("layout") or "").strip() or None,
        notes=str(row.get("notes") or "").strip() or None,
        raw=row,
    )


def normalize_recipient(row: dict[str, Any]) -> Recipient:
    return Recipient(
        email=str(row.get("email") or "").strip().lower(),
        name=str(row.get("name") or row.get("firstname") or "").strip() or None,
        consent=_parse_bool(row.get("consent"), True),
        segment=str(row.get("segment") or "").strip() or None,
    )


def load_properties_from_path(path: Path) -> list[Property]:
    return [normalize_property(row) for row in read_rows(path)]


def load_recipients_from_path(path: Path | None) -> list[Recipient]:
    if path is None or not path.exists():
        return []
    return [normalize_recipient(row) for row in read_rows(path)]
